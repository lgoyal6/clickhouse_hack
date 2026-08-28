"""Convert an OFLC .xlsx disclosure file to CSV once, then load from CSV.

    python -m ingest.convert data/LCA_Disclosure_Data_FY2025_Q4.xlsx

Reloads during a build happen more often than anyone plans for, and parsing a
75 MB spreadsheet takes minutes while the same data as CSV takes seconds. Convert
once, commit nothing (data/ is gitignored), reload freely.

Streams with openpyxl read_only so a 75 MB file does not become a 2 GB process.
"""
from __future__ import annotations

import argparse
import csv
import pathlib
import sys
import time


def convert(src: pathlib.Path, dst: pathlib.Path | None = None) -> pathlib.Path:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("pip install openpyxl") from None

    dst = dst or src.with_suffix(".csv")
    started = time.time()

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_out = 0
    blanks = 0
    with dst.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [str(c) if c is not None else "" for c in row]
                writer.writerow(header)
                print(f"{len(header)} columns")
                continue
            # Skip trailing blank rows. openpyxl iterates the sheet's declared
            # dimension, and the real FY2025 Q4 file declares 563,689 rows while
            # holding 118,580 of data. Without this the loader sees 445,109 empty
            # rows and every distribution is computed over mostly nothing.
            if not any(c is not None and str(c).strip() for c in row):
                blanks += 1
                continue
            writer.writerow(["" if c is None else c for c in row])
            rows_out += 1
            if rows_out % 100_000 == 0:
                print(f"  {rows_out:,} rows ({time.time() - started:.0f}s)", flush=True)
    wb.close()

    print(f"{rows_out:,} rows -> {dst} in {time.time() - started:.0f}s "
          f"({dst.stat().st_size / 1048576:.0f} MB); {blanks:,} blank rows skipped")
    return dst


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("path", type=pathlib.Path)
    ap.add_argument("--out", type=pathlib.Path, default=None)
    args = ap.parse_args(argv)
    if not args.path.exists():
        raise SystemExit(f"{args.path} does not exist")
    convert(args.path, args.out)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
