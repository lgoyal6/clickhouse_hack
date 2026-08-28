"""Print the actual column headers of an OFLC file.

    python -m ingest.headers path/to/file.xlsx

Run this BEFORE trusting anything in column_maps.py. The single most likely way this
project ships a wrong number is a mismapped wage column.
"""
import sys
import pathlib


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print(__doc__)
        return 2
    path = pathlib.Path(argv[1])
    if path.suffix.lower() in {".csv", ".txt"}:
        import csv
        with path.open(newline="", encoding="utf-8", errors="replace") as fh:
            header = next(csv.reader(fh))
    elif path.suffix.lower() in {".xlsx", ".xls"}:
        try:
            import openpyxl
        except ImportError:
            print("pip install openpyxl")
            return 2
        wb = openpyxl.load_workbook(path, read_only=True)
        header = [c.value for c in next(wb[wb.sheetnames[0]].iter_rows(max_row=1))]
    else:
        print(f"unsupported file type: {path.suffix}")
        return 2

    print(f"{len(header)} columns in {path.name}:\n")
    for i, name in enumerate(header):
        print(f"  {i:3d}  {name}")

    from .column_maps import LCA_MODERN
    wanted = {v for v in LCA_MODERN.values() if v}
    missing = sorted(wanted - set(header))
    if missing:
        print("\nEXPECTED BUT NOT PRESENT (fix column_maps.py before loading):")
        for name in missing:
            print(f"  {name}")
    else:
        print("\nEvery column LCA_MODERN expects is present.")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
